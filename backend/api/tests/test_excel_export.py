"""Tests for T16: Excel export."""
import os
import tempfile
from datetime import timedelta
from unittest.mock import MagicMock, patch

from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from api.models import Bus, Office, Settlement, Trip, User
from api.tasks import generate_excel_report


class ExcelExportTests(TestCase):
    """Export trigger, status polling, download, and access control."""

    def setUp(self):
        self.office = Office.objects.create(name='HQ', city='Algiers')
        self.office_b = Office.objects.create(name='Branch', city='Oran')
        self.admin = User.objects.create_superuser(
            phone='0550000001', password='pass',
            first_name='Admin', last_name='X',
        )
        self.conductor = User.objects.create_user(
            phone='0550000003', password='pass',
            first_name='Cond', last_name='A',
            role='conductor', office=self.office,
        )
        self.bus = Bus.objects.create(
            plate_number='EXP-001', capacity=50, office=self.office,
        )
        self.client = APIClient()

    @patch('api.views.export_views.generate_excel_report')
    def test_trigger_export(self, mock_task):
        """POST /exports/ triggers Celery task and returns task_id."""
        mock_result = MagicMock()
        mock_result.id = 'test-task-id-123'
        mock_task.delay.return_value = mock_result

        self.client.force_authenticate(self.admin)
        resp = self.client.post('/api/exports/', {
            'report_type': 'daily',
            'filters': {'date_from': '2026-02-01', 'date_to': '2026-02-18'},
        }, format='json')
        self.assertEqual(resp.status_code, status.HTTP_202_ACCEPTED)
        self.assertEqual(resp.data['task_id'], 'test-task-id-123')
        self.assertIn('download_token', resp.data)

    def test_invalid_report_type(self):
        """Reject invalid report type."""
        self.client.force_authenticate(self.admin)
        resp = self.client.post('/api/exports/', {
            'report_type': 'invalid',
        }, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_conductor_cannot_export(self):
        """Conductors are denied export access."""
        self.client.force_authenticate(self.conductor)
        resp = self.client.post('/api/exports/', {
            'report_type': 'daily',
        }, format='json')
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_download_invalid_token(self):
        """Download with invalid token is rejected."""
        self.client.force_authenticate(self.admin)
        resp = self.client.get('/api/exports/fake-task/download/?token=invalidtoken')
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    @patch('celery.result.AsyncResult')
    def test_export_status_is_lowercase(self, mock_async_result):
        result = MagicMock()
        result.status = 'SUCCESS'
        result.ready.return_value = True
        result.successful.return_value = True
        result.failed.return_value = False
        result.result = {'filename': 'daily_test.xlsx'}
        mock_async_result.return_value = result

        self.client.force_authenticate(self.admin)
        resp = self.client.get('/api/exports/task-123/status/')

        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data['status'], 'success')
        self.assertEqual(resp.data['filename'], 'daily_test.xlsx')

    @override_settings(MEDIA_ROOT=tempfile.gettempdir())
    def test_generated_workbook_includes_settlement_sheet_and_trip_columns(self):
        trip = Trip.objects.create(
            origin_office=self.office,
            destination_office=self.office_b,
            conductor=self.conductor,
            bus=self.bus,
            departure_datetime=timezone.now() - timedelta(hours=4),
            arrival_datetime=timezone.now() - timedelta(hours=2),
            status='completed',
            passenger_base_price=1000,
            cargo_small_price=500,
            cargo_medium_price=1000,
            cargo_large_price=1500,
        )
        Settlement.objects.create(
            trip=trip,
            office=self.office_b,
            conductor=self.conductor,
            expected_passenger_cash=1200,
            expected_cargo_cash=800,
            expected_total_cash=2000,
            agency_presale_total=300,
            outstanding_cargo_delivery=400,
            expenses_to_reimburse=150,
            net_cash_expected=1850,
            actual_cash_received=2000,
            actual_expenses_reimbursed=150,
            discrepancy_amount=0,
            status='settled',
            calculation_snapshot={'expected_total_cash': 2000},
            settled_at=timezone.now() - timedelta(hours=1),
        )

        result = generate_excel_report.run(
            'daily',
            {
                'date_from': str(timezone.localdate() - timedelta(days=1)),
                'date_to': str(timezone.localdate() + timedelta(days=1)),
            },
            self.admin.id,
        )

        filepath = os.path.join(tempfile.gettempdir(), 'exports', result['filename'])
        self.assertTrue(os.path.exists(filepath))

        workbook = load_workbook(filepath)
        self.assertIn('Settlements', workbook.sheetnames)

        trip_headers = [cell.value for cell in workbook['Trips'][1]]
        self.assertIn('Settlement Status', trip_headers)
        self.assertIn('Settlement Discrepancy', trip_headers)

        settlement_headers = [cell.value for cell in workbook['Settlements'][1]]
        self.assertIn('Expected Total Cash', settlement_headers)
        self.assertIn('Discrepancy', settlement_headers)
