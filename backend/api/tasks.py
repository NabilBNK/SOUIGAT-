import logging
import os
import uuid
from datetime import date, timedelta

from celery import shared_task
from django.conf import settings
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task
def cleanup_expired_tokens():
    """Delete expired outstanding tokens. Runs daily via Celery Beat."""
    from rest_framework_simplejwt.token_blacklist.models import OutstandingToken

    cutoff = timezone.now() - timedelta(days=30)
    deleted, _ = OutstandingToken.objects.filter(expires_at__lt=cutoff).delete()
    if deleted:
        logger.info("Cleaned up %d expired tokens", deleted)


@shared_task(bind=True, max_retries=2, soft_time_limit=120)
def generate_excel_report(self, report_type, filters, user_id):
    """
    Generate Excel export as .xlsx file.

    Sheets per report_type:
    - daily: Trips, Tickets, Cargo, Expenses
    - trip: Single trip breakdown
    - route: Route aggregation
    """
    from openpyxl import Workbook
    from django.db.models import Sum, Value
    from django.db.models.functions import Coalesce

    from api.models import (
        CargoTicket, PassengerTicket, Settlement, Trip, TripExpense, User,
    )

    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return {'error': 'User not found'}

    wb = Workbook()

    date_from = filters.get('date_from')
    date_to = filters.get('date_to')
    office_id = filters.get('office_id')

    try:
        date_from = date.fromisoformat(date_from) if date_from else date.today()
        date_to = date.fromisoformat(date_to) if date_to else date.today()
    except (ValueError, TypeError):
        date_from = date.today()
        date_to = date.today()

    # Scope trips
    trips = Trip.objects.filter(
        departure_datetime__date__gte=date_from,
        departure_datetime__date__lte=date_to,
    )
    if office_id:
        from django.db.models import Q
        trips = trips.filter(
            Q(origin_office_id=office_id) | Q(destination_office_id=office_id)
        )
    elif user.role == 'office_staff':
        from django.db.models import Q
        trips = trips.filter(
            Q(origin_office_id=user.office_id)
            | Q(destination_office_id=user.office_id)
        )

    trip_ids = list(trips.values_list('id', flat=True))

    # --- Trips sheet ---
    ws_trips = wb.active
    ws_trips.title = 'Trips'
    ws_trips.append([
        'ID', 'Origin', 'Destination', 'Departure', 'Status',
        'Conductor', 'Base Price', 'Settlement Status', 'Settlement Discrepancy',
    ])
    settlements_by_trip = {
        settlement.trip_id: settlement
        for settlement in Settlement.objects.filter(trip_id__in=trip_ids)
    }
    for t in trips.select_related('origin_office', 'destination_office', 'conductor'):
        settlement = settlements_by_trip.get(t.id)
        ws_trips.append([
            t.id,
            str(t.origin_office),
            str(t.destination_office),
            t.departure_datetime.strftime('%Y-%m-%d %H:%M'),
            t.status,
            str(t.conductor) if t.conductor else '',
            t.passenger_base_price,
            settlement.status if settlement else '',
            settlement.discrepancy_amount if settlement and settlement.discrepancy_amount is not None else '',
        ])

    # --- Tickets sheet ---
    ws_tickets = wb.create_sheet('Passenger Tickets')
    ws_tickets.append([
        'ID', 'Trip', 'Ticket #', 'Passenger', 'Price',
        'Currency', 'Payment', 'Seat', 'Status',
    ])
    for tk in PassengerTicket.objects.filter(trip_id__in=trip_ids).select_related('trip'):
        ws_tickets.append([
            tk.id, tk.trip_id, tk.ticket_number,
            tk.passenger_name, tk.price, tk.currency,
            tk.payment_source, tk.seat_number, tk.status,
        ])

    # --- Cargo sheet ---
    ws_cargo = wb.create_sheet('Cargo Tickets')
    ws_cargo.append([
        'ID', 'Trip', 'Ticket #', 'Sender', 'Receiver',
        'Tier', 'Price', 'Currency', 'Status',
    ])
    for c in CargoTicket.objects.filter(trip_id__in=trip_ids).select_related('trip'):
        ws_cargo.append([
            c.id, c.trip_id, c.ticket_number,
            c.sender_name, c.receiver_name,
            c.cargo_tier, c.price, c.currency, c.status,
        ])

    # --- Expenses sheet ---
    ws_exp = wb.create_sheet('Expenses')
    ws_exp.append(['ID', 'Trip', 'Description', 'Amount', 'Currency'])
    for e in TripExpense.objects.filter(trip_id__in=trip_ids):
        ws_exp.append([
            e.id, e.trip_id, e.description, e.amount, e.currency,
        ])

    # --- Settlements sheet ---
    ws_settle = wb.create_sheet('Settlements')
    ws_settle.append([
        'ID', 'Trip', 'Office', 'Conductor', 'Status',
        'Expected Passenger Cash', 'Expected Cargo Cash', 'Expected Total Cash',
        'Agency Presale', 'Outstanding POD', 'Expenses Reimbursed Expected',
        'Net Cash Expected', 'Actual Cash Received', 'Actual Expenses Reimbursed',
        'Discrepancy', 'Created At', 'Settled At', 'Disputed At', 'Resolved At',
    ])
    for settlement in Settlement.objects.filter(trip_id__in=trip_ids).select_related(
        'office', 'conductor',
    ).order_by('trip_id'):
        ws_settle.append([
            settlement.id,
            settlement.trip_id,
            str(settlement.office),
            str(settlement.conductor),
            settlement.status,
            settlement.expected_passenger_cash,
            settlement.expected_cargo_cash,
            settlement.expected_total_cash,
            settlement.agency_presale_total,
            settlement.outstanding_cargo_delivery,
            settlement.expenses_to_reimburse,
            settlement.net_cash_expected,
            settlement.actual_cash_received if settlement.actual_cash_received is not None else '',
            settlement.actual_expenses_reimbursed,
            settlement.discrepancy_amount if settlement.discrepancy_amount is not None else '',
            settlement.created_at.strftime('%Y-%m-%d %H:%M') if settlement.created_at else '',
            settlement.settled_at.strftime('%Y-%m-%d %H:%M') if settlement.settled_at else '',
            settlement.disputed_at.strftime('%Y-%m-%d %H:%M') if settlement.disputed_at else '',
            settlement.resolved_at.strftime('%Y-%m-%d %H:%M') if settlement.resolved_at else '',
        ])

    # Save file
    export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
    os.makedirs(export_dir, exist_ok=True)
    filename = f'{report_type}_{date_from}_{uuid.uuid4().hex[:8]}.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    logger.info('Export generated: %s (%d trips)', filename, len(trip_ids))
    return {'filename': filename, 'trip_count': len(trip_ids)}


@shared_task
def cleanup_old_exports():
    """Delete export files older than 7 days. Runs daily via Celery Beat."""
    export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
    if not os.path.exists(export_dir):
        return 'No exports directory.'

    cutoff = timezone.now().timestamp() - (7 * 86400)
    deleted = 0
    for f in os.listdir(export_dir):
        filepath = os.path.join(export_dir, f)
        if os.path.isfile(filepath) and os.path.getmtime(filepath) < cutoff:
            os.remove(filepath)
            deleted += 1

    logger.info('Cleaned up %d old export files', deleted)
    return f'Deleted {deleted} files'


@shared_task
def cleanup_old_synclogs():
    """Delete SyncLog entries older than 30 days. Runs weekly."""
    from api.models import SyncLog

    cutoff = timezone.now() - timedelta(days=30)
    deleted, _ = SyncLog.objects.filter(created_at__lt=cutoff).delete()
    if deleted:
        logger.info('Cleaned up %d old sync logs', deleted)
    return f'Deleted {deleted} sync logs'
